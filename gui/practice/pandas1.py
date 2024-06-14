## 목표 : 엑셀 파일에서 1행 1열의 데이터를 출력한다 ##
import pandas as pd
from openpyxl import load_workbook

excel_file = 'sample.xlsx'

# 엑셀 파일 열기
workbook = load_workbook(filename=excel_file)

# 첫 번째 시트 선택
sheet = workbook.active

# 시트에서 데이터 읽어오기
for row in sheet.iter_rows(values_only=True):
    print(row)

# 엑셀 파일 닫기
workbook.close()